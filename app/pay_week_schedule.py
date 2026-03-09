import tkinter as tk
from tkinter import messagebox, Toplevel, Label, Button
from session_context import verify_admin, exit_session, enforce_plan
import os
from datetime import datetime, timedelta
from collections import defaultdict
from openpyxl import Workbook, load_workbook
import platform
import re
from Admin_Interface import AdminInterfaceFrame
from startup_page import StartPageFrame

from db_connection import get_db_connection

def weekday_to_date(day_name, reference_date=None, force_next_week=False):
    weekday_map = {
        "Monday": 0, "Tuesday": 1, "Wednesday": 2,
        "Thursday": 3, "Friday": 4, "Saturday": 5, "Sunday": 6
    }
    if not reference_date:
        reference_date = datetime.today()
    target_day = weekday_map[day_name.capitalize()]
    days_ahead = (target_day - reference_date.weekday()) % 7
    if force_next_week:
        days_ahead += 7 if days_ahead == 0 else 0
    return reference_date + timedelta(days=days_ahead)


def submit_data(start_day, end_day):
    cnxn = get_db_connection()
    with cnxn as conn:
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS pay_week (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                start_day TEXT,
                end_day TEXT
            )
        """)
        cursor.execute(
            "INSERT INTO pay_week (start_day, end_day) VALUES (?, ?)",
            (start_day.capitalize(), end_day.capitalize())
        )
        conn.commit()
    messagebox.showinfo("Saved", f"Pay week:\nStart: {start_day}\nEnd: {end_day}")


def fetch_payroll_data(start_date, end_date):
    cnxn = get_db_connection()
    with cnxn as conn:
        cursor = conn.cursor()
        cursor.execute(
            """SELECT MM_DD_YYYY, First_Name, Last_Name, Day_of_Week,
                      Employee_ID, Employee_Hours, Employee_Job_Title,
                      Builder, Jobsite, Model_Name, Lot_Number, Block_Number,
                      total_pay, Pay_Per_Employee, Split_Pay_Per_Employee
               FROM AIAI_Weekly_Payroll
               WHERE MM_DD_YYYY BETWEEN ? AND ?""",
            (start_date.strftime("%m/%d/%Y"), end_date.strftime("%m/%d/%Y"))
        )
        records = cursor.fetchall()

    grouped = defaultdict(list)
    for row in records:
        fname, lname, emp_id = row[1], row[2], row[4]
        key = f"{fname}_{lname}" if fname and lname else f"Employee_{emp_id}"
        grouped[key].append(row)
    return grouped


def export_to_excel(grouped_data, start_date, end_date):
    folder = end_date.strftime("%Y-%m")
    export_path = os.path.join(
        os.path.expanduser("~"),
        "Documents",
        "AIAI_PM_App",
        "Payroll_Excels",
        folder
    )
    os.makedirs(export_path, exist_ok=True)
    workbook_path = os.path.join(export_path, f"Payroll_{folder}.xlsx")

    workbook = load_workbook(workbook_path) if os.path.exists(workbook_path) else Workbook()
    if "Sheet" in workbook.sheetnames:
        del workbook["Sheet"]

    week_label = f"Week {((end_date.day - 1) // 7) + 1}"

    for key, rows in grouped_data.items():
        sanitized = re.sub(r"[\\/*?:[\]]", "_", key)
        sheet_name = re.sub(r"[^\w\s\-]", "_", f"{sanitized}_{week_label}")[:31]
        if sheet_name in workbook.sheetnames:
            continue

        ws = workbook.create_sheet(sheet_name)
        headers = [
            "MM_DD_YYYY", "First_Name", "Last_Name", "Day_of_Week",
            "Employee_ID", "Employee_Hours", "Employee_Job_Title",
            "Builder", "Jobsite", "Model_Name", "Lot_Number", "Block_Number",
            "total_pay", "Pay_Per_Employee", "Split_Pay_Per_Employee"
        ]
        ws.append(headers)

        total_hours = 0
        total_pay = 0
        for row in rows:
            ws.append(row)
            try:
                total_hours += float(row[5])
                total_pay += float(row[12])
            except Exception:
                pass

        ws.append([""] * 12 + ["Total:", total_hours, total_pay])

    workbook.save(workbook_path)
    show_export_popup(workbook_path)


def show_export_popup(path):
    popup = Toplevel()
    popup.title("Export Complete")
    popup.geometry("400x150")
    popup.resizable(False, False)

    Label(
        popup,
        text="✅ Payroll workbook created.\nDo you want to open it?",
        font=("Arial", 11)
    ).pack(pady=20)

    def open_excel():
        try:
            if platform.system() == "Windows":
                os.startfile(path)
            else:
                import subprocess
                subprocess.call(["open", path])
        except Exception as e:
            messagebox.showerror("Error", f"Could not open file.\n{e}")
        popup.destroy()

    Button(popup, text="Open Workbook", command=open_excel, width=20).pack(pady=5)
    Button(popup, text="Close", command=popup.destroy, width=20).pack(pady=5)


def get_latest_pay_week():
    cnxn = get_db_connection()
    with cnxn as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT start_day, end_day FROM pay_week ORDER BY id DESC LIMIT 1")
        result = cursor.fetchone()
    return result if result and all(result) else None


def conditional_run_export():
    entry = get_latest_pay_week()
    if not entry:
        return
    start_day, end_day = entry
    end_date = weekday_to_date(end_day, force_next_week=True).date()
    start_date = weekday_to_date(start_day).date()
    if datetime.today().date() == end_date + timedelta(days=2):
        grouped = fetch_payroll_data(start_date, end_date)
        export_to_excel(grouped, start_date, end_date)


class PayWeekScheduleFrame(tk.Frame):
    def __init__(self, controller):
        super().__init__(controller)
        self.controller = controller
        self.start_var = tk.StringVar()
        self.end_var = tk.StringVar()
        self.build_ui()

    def on_show(self):
        # ADMIN CHECK
        if not verify_admin():
            messagebox.showerror("User not authorized")
            exit_session()
            return False

        # PLAN CHECK
        if not enforce_plan("pro", "enterprise"):
            messagebox.showerror("Access Denied", "You must be subscribed to a plan.")
            self.controller.show_frame(StartPageFrame)
            return False

        return True    

    def build_ui(self):
        self.configure(bg="white")

        tk.Label(
            self,
            text=(
                "Enter the start and end day for the pay period.\n"
                "Start Day: First day of employee work.\n"
                "End Day: Last day of pay week.\n"
                "Payroll will be exported 2 days after the end day."
            ),
            wraplength=900,
            justify="left",
            font=("Arial", 12),
            fg="darkblue",
            bg="white"
        ).grid(row=0, column=0, columnspan=2, padx=10, pady=20, sticky="w")

        tk.Label(self, text="Start Day (e.g. Thursday):", bg="white").grid(
            row=1, column=0, padx=5, pady=5, sticky="e"
        )
        tk.Entry(self, textvariable=self.start_var).grid(
            row=1, column=1, padx=5, pady=5
        )

        tk.Label(self, text="End Day (e.g. Wednesday):", bg="white").grid(
            row=2, column=0, padx=5, pady=5, sticky="e"
        )
        tk.Entry(self, textvariable=self.end_var).grid(
            row=2, column=1, padx=5, pady=5
        )

        tk.Button(self, text="Submit", command=self.on_submit, bg="green").grid(
            row=3, column=0, columnspan=2, pady=20
        )
        tk.Button(
            self,
            text="Admin Interface",
            command=lambda: self.controller.show_frame(AdminInterfaceFrame), bg="gold"
        ).grid(row=4, column=0, columnspan=2, pady=20)

    def on_submit(self):
        start = self.start_var.get().strip()
        end = self.end_var.get().strip()
        if not start or not end:
            messagebox.showwarning("Invalid Input", "Please enter both start and end days.")
            return
        submit_data(start, end)
        self.controller.show_frame(AdminInterfaceFrame)


def main():
    pass