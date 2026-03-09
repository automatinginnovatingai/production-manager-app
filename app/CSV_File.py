import os
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook, load_workbook
from datetime import datetime
from Admin_Interface import AdminInterfaceFrame
from session_context import verify_admin, enforce_plan, exit_session
from db_connection import get_db_connection


def normalize(val):
    return str(val).strip().lower() if val is not None else ""


def export_to_excel():
    # ADMIN CHECK
    if not verify_admin():
        messagebox.showerror("Unauthorized", "You do not have permission to export payroll data.")
        exit_session()
        return

    # PLAN CHECK
    if not enforce_plan("pro", "enterprise"):
        messagebox.showerror(
            "Access Denied",
            "Your plan does not include access to this feature. Upgrade plan."
        )
        AdminInterfaceFrame()
        return

    suggested_name = "Weekly_Payroll.xlsx"
    save_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        title="Save Excel file",
        initialfile=suggested_name
    )
    if not save_path:
        messagebox.showinfo("Info", "Excel export cancelled.")
        return

    current_date = datetime.now()
    day = current_date.strftime("%d")
    month = current_date.strftime("%B %Y")

    # SQL SERVER CONNECTION
    conn = get_db_connection()
    cursor = conn.cursor()

    # Fetch all payroll rows
    cursor.execute("SELECT * FROM AIAI_Weekly_Payroll")
    data = cursor.fetchall()

    # Get SQL Server column names
    columns = [desc[0] for desc in cursor.description]
    column_indices = {column: index for index, column in enumerate(columns)}

    # Define Excel export order
    column_order = [
        'Time', 'MM_DD_YYYY', 'Day_of_Week', 'First_Name', 'Last_Name', 'Employee_ID',
        'Employee_Hours', 'Employee_Job_Title', 'Builder', 'Jobsite', 'Model_Name',
        'Lot_Number', 'Block_Number', 'Material_Used', 'R_Value', 'Material_Width',
        'Sqft_Bags_Installed', 'Pay_Per_Tube_Piece_Rate', 'Pay', 'Material_Used_2',
        'R_Value_2', 'Material_Width_2', 'Sqft_Bags_Installed_2', 'Pay_Per_Tube_Piece_Rate_2',
        'Pay_2', 'Material_Used_3', 'R_Value_3', 'Material_Width_3', 'Sqft_Bags_Installed_3',
        'Pay_Per_Tube_Piece_Rate_3', 'Pay_3', 'Material_Used_4', 'R_Value_4', 'Material_Width_4',
        'Sqft_Bags_Installed_4', 'Pay_Per_Tube_Piece_Rate_4', 'Pay_4', 'Material_Used_5',
        'R_Value_5', 'Material_Width_5', 'Sqft_Bags_Installed_5', 'Pay_Per_Tube_Piece_Rate_5',
        'Pay_5', 'Material_Used_6', 'R_Value_6', 'Material_Width_6', 'Sqft_Bags_Installed_6',
        'Pay_Per_Tube_Piece_Rate_6', 'Pay_6', 'Material_Used_7', 'R_Value_7', 'Material_Width_7',
        'Sqft_Bags_Installed_7', 'Pay_Per_Tube_Piece_Rate_7', 'Pay_7', 'total_pay',
        'Pay_Per_Employee', 'Split_Pay_Per_Employee'
    ]

    # Reorder rows to match Excel column order
    reordered_data = []
    for row in data:

        # Validate critical fields (now plain text)
        valid = True
        for field in ['First_Name', 'Last_Name', 'Employee_Hours', 'Employee_ID']:
            idx = column_indices[field]
            val = row[idx]
            if not str(val).strip():
                valid = False
                break

        if not valid:
            continue

        reordered_row = [row[column_indices[col]] for col in column_order]
        reordered_data.append(reordered_row)

    # Load or create Excel file
    if os.path.exists(save_path):
        workbook = load_workbook(save_path)
    else:
        workbook = Workbook()
        del workbook["Sheet"]

    # Month sheet
    if month not in workbook.sheetnames:
        workbook.create_sheet(title=month)

    # Day sheet
    day_sheet_name = f"{month} - {day}"
    if day_sheet_name not in workbook.sheetnames:
        day_sheet = workbook.create_sheet(title=day_sheet_name)
    else:
        day_sheet = workbook[day_sheet_name]

    # Write header if empty
    if day_sheet.max_row == 1:
        day_sheet.append(column_order)

    # Avoid duplicates
    existing_entries = set()
    for existing_row in day_sheet.iter_rows(min_row=2, values_only=True):
        key = tuple(normalize(existing_row[column_order.index(c)]) for c in [
            'First_Name', 'Last_Name', 'MM_DD_YYYY', 'Builder', 'Jobsite', 'Lot_Number'
        ])
        existing_entries.add(key)

    new_rows_added = 0
    for row in reordered_data:
        key = tuple(normalize(row[column_order.index(c)]) for c in [
            'First_Name', 'Last_Name', 'MM_DD_YYYY', 'Builder', 'Jobsite', 'Lot_Number'
        ])
        if key not in existing_entries:
            day_sheet.append(row)
            existing_entries.add(key)
            new_rows_added += 1

    # Footer timestamp
    if new_rows_added:
        timestamp = datetime.now().strftime("Exported on %Y-%m-%d at %H:%M")
        footer_cell = f"A{day_sheet.max_row + 2}"
        day_sheet[footer_cell] = timestamp
        day_sheet.row_dimensions[day_sheet.max_row + 2].hidden = True
        messagebox.showinfo("Success", f"{new_rows_added} new rows added to {day_sheet_name}")
    else:
        messagebox.showinfo("Info", f"No new entries found. Nothing changed in {day_sheet_name}")

    # Save and open file
    try:
        os.startfile(save_path)
    except Exception as e:
        messagebox.showerror("Error", f"Could not open the file automatically.\n{e}")

    workbook.save(save_path)
    conn.close()


if __name__ == '__main__':
    export_to_excel()